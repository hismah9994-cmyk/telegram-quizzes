"""
Daily Telegram Quiz Bot
Posts 4 daily quizzes (Physics, Chemistry, Biology, Math) based on the Excel plan.
Each quiz is sent as an image + quiz poll with the correct answer marked.
"""
import os
import sys
import json
import datetime
from pathlib import Path

import openpyxl
import requests

# ------------------------------------------------------------------
# Configuration (read from environment so we can deploy to any runner)
# ------------------------------------------------------------------
ROOT = Path(__file__).parent

TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "").strip()
# Day N of the 350-day plan that corresponds to START_DATE (Asia/Riyadh)
START_DAY = int(os.environ.get("START_DAY", "40"))
START_DATE = os.environ.get("START_DATE", "2026-04-16")

# Timezone (KSA / Iraq / Yemen / Kuwait)
TZ = datetime.timezone(datetime.timedelta(hours=3))

SUBJECTS = [
    ("phys", "Physics"),
    ("chem", "Chemistry"),
    ("bio",  "Biology"),
    ("math", "Math"),
]


# ------------------------------------------------------------------
# Core helpers
# ------------------------------------------------------------------
def today_day_number() -> int:
    """Compute today's day-of-plan based on START_DATE/START_DAY in KSA TZ."""
    today = datetime.datetime.now(TZ).date()
    start = datetime.date.fromisoformat(START_DATE)
    return START_DAY + (today - start).days


def load_plan() -> dict:
    xlsx = next(ROOT.glob("*.xlsx"))
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    plan = {}
    for r in range(2, ws.max_row + 1):
        day = ws.cell(r, 1).value
        if day is None:
            continue
        plan[int(day)] = {
            "math":   int(ws.cell(r, 2).value),
            "bio":    int(ws.cell(r, 3).value),
            "chem":   int(ws.cell(r, 4).value),
            "phys":   int(ws.cell(r, 5).value),
            "module": int(ws.cell(r, 6).value),
        }
    return plan


def load_answers() -> dict:
    with open(ROOT / "_build" / "answers.json", "r", encoding="utf-8") as f:
        return json.load(f)


# ------------------------------------------------------------------
# Telegram API
# ------------------------------------------------------------------
def tg(method: str, **payload):
    url = f"https://api.telegram.org/bot{TOKEN}/{method}"
    files = payload.pop("_files", None)
    r = requests.post(url, data=payload, files=files, timeout=60)
    if not r.ok:
        raise RuntimeError(f"Telegram {method} failed: {r.status_code} {r.text}")
    data = r.json()
    if not data.get("ok"):
        raise RuntimeError(f"Telegram {method} not ok: {data}")
    return data["result"]


def send_message(text):
    return tg("sendMessage", chat_id=CHAT_ID, text=text, parse_mode="HTML")


def send_photo(image_path, caption=None):
    with open(image_path, "rb") as f:
        return tg(
            "sendPhoto",
            chat_id=CHAT_ID,
            caption=caption or "",
            _files={"photo": f},
        )


def send_quiz_poll(question, correct_index):
    return tg(
        "sendPoll",
        chat_id=CHAT_ID,
        question=question,
        options=json.dumps(["A", "B", "C", "D"], ensure_ascii=False),
        type="quiz",
        correct_option_id=correct_index,
        # Anonymous: no one (not even admins) can see who voted for what.
        is_anonymous="true",
    )


# ------------------------------------------------------------------
# Main workflow
# ------------------------------------------------------------------
def run():
    if not TOKEN or not CHAT_ID:
        sys.exit("ERROR: TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID must be set.")

    plan = load_plan()
    answers = load_answers()
    day = today_day_number()

    print(f"[info] Day {day}")

    if day not in plan:
        print(f"[warn] No plan for day {day} — nothing to send.")
        return

    p = plan[day]
    mod = p["module"]
    module_ans = answers[str(mod)]

    for key, name_en in SUBJECTS:
        qn = p[key]
        letter = module_ans[qn - 1]
        idx = "ABCD".index(letter)
        img = ROOT / "35 modules" / str(mod) / f"{qn}.png"
        if not img.exists():
            print(f"[warn] Missing image: {img}")
            continue
        send_photo(str(img), caption=name_en)
        send_quiz_poll(f"{name_en} - choose answer", idx)
        print(f"[ok] {key} Q{qn} -> {letter}")

    print("[done] quiz posted successfully")


if __name__ == "__main__":
    run()
